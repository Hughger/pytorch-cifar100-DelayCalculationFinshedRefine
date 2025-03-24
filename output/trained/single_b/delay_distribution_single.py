import pandas as pd
import numpy as np
import argparse
import openpyxl

def write_to_excel(sum_results, count_results, opt_results, excel_path):
    """
    将统计结果写入 Excel 表格，组织成两行数据。

    Args:
        sum_results (list): 包含每个子表求和相乘的结果。
        count_results (list): 包含每个子表元素总数相乘的结果。
        excel_path (str): Excel 文件的路径。
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    # 写入表头
    sheet.cell(row=1, column=1).value = "Type"
    for i in range(len(sum_results)):
      sheet.cell(row=1, column=i+2).value = f"Subtable_{i+1}"

    # 写入数据 - 第一行：求和结果
    sheet.cell(row=2, column=1).value = "Sum_Multiply"
    for i, value in enumerate(sum_results):
       sheet.cell(row=2, column=i + 2).value = value

    # 写入数据 - 第二行：元素总数结果
    sheet.cell(row=3, column=1).value = "Count_Multiply"
    for i, value in enumerate(count_results):
        sheet.cell(row=3, column=i + 2).value = value

    # 写入数据 - 第三行：优化的比例
    sheet.cell(row=4, column=1).value = "Delay_Optimization"
    for i, value in enumerate(opt_results):
        sheet.cell(row=4, column=i + 2).value = value

    workbook.save(excel_path)

def process_data(excel_file, text_file, output_file):
    """
    读取 Excel 文件和文本文件，对每个子表求和并乘以相应的文本行数据，然后将结果向量写入 Excel 文件。

    Args:
        excel_file (str): Excel 文件路径。
        text_file (str): 文本文件路径。
        output_file (str): 输出 Excel 文件路径。
    """
    # 读取 Excel 文件
    xls = pd.ExcelFile(excel_file)
    sheet_names = xls.sheet_names

    # 读取文本文件
    with open(text_file, 'r') as f:
        text_data = [float(line.strip()) for line in f]

    if len(sheet_names) != len(text_data):
        print(f"sheet_num = {len(sheet_names)}")
        print(f"txt_num = {len(text_data)}")
        raise ValueError("子表数量与文本文件数据行数不匹配")

    results = []
    counts = []
    opts = []

    # 遍历子表
    for i, sheet_name in enumerate(sheet_names):
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        # 计算子表总和
        sum_of_data = df.values.sum()
        # 计算子表元素总数
        count_of_data = df.size
        # 将子表总和乘以文本数据
        result = sum_of_data * text_data[i]
        # 将元素总数乘以文本数据
        count = count_of_data * text_data[i]
        # 求优化比例
        opt = 1 - result / count

        results.append(result)
        counts.append(count)
        opts.append(opt)

    # 使用 openpyxl 写入 Excel 文件
    write_to_excel(results, counts, opts, output_file)
    print(f"结果向量已写入到：{output_file}")


# 示例用法
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-net', type=str, default='vgg16', help='net type')
    args = parser.parse_args()
    excel_file = f"{args.net}/single_matrix_tmp.xlsx"
    txt_file = f"./../compute_repeat_element_{args.net}.txt"

    output_file = f"{args.net}/single_delay_distribution.xlsx"  # 输出Excel文件
    try:
        process_data(excel_file, txt_file, output_file)
    except FileNotFoundError:
         print(f"错误：文件未找到，请检查输入文件路径是否正确。")
    except ValueError as e:
        print(f"错误: {e}")
    except Exception as e:
        print(f"发生未知错误：{e}")